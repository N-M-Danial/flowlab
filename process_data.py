"""
pipeline/process_data.py
=========================
Contains all data-processing logic extracted from the original generate_map.py.
Called by app.py to produce JSON that the frontend consumes.
"""

import json
import os

try:
    import openpyxl
except ImportError:
    raise ImportError(
        "Missing package: openpyxl\n  Install with: pip install openpyxl"
    )


def compute_los(vc):
    if vc <= 0.60: return "A"
    elif vc <= 0.70: return "B"
    elif vc <= 0.80: return "C"
    elif vc <= 0.90: return "D"
    elif vc <= 1.00: return "E"
    else: return "F"


def run(geojson_path, xlsx_path):
    """
    Process the GeoJSON and Excel files and return:
      (traffic_data, road_coords, map_bounds, date_label, road_count)

    Raises FileNotFoundError if either input file is missing.
    """

    # ── Load road geometries ──────────────────────────────────────────────────
    if not os.path.exists(geojson_path):
        raise FileNotFoundError(
            f"Road geometry file not found: {geojson_path}\n"
            "Please upload road_geometries.geojson via the dashboard."
        )

    with open(geojson_path, encoding="utf-8") as f:
        geojson = json.load(f)

    road_coords = {}
    all_lngs, all_lats = [], []

    for feature in geojson["features"]:
        name = feature["properties"]["road_name"]
        geom = feature["geometry"]
        if geom["type"] == "LineString":
            road_coords[name] = geom["coordinates"]
            for c in geom["coordinates"]:
                all_lngs.append(c[0])
                all_lats.append(c[1])
        elif geom["type"] == "MultiLineString":
            road_coords[name] = geom["coordinates"]
            for seg in geom["coordinates"]:
                for c in seg:
                    all_lngs.append(c[0])
                    all_lats.append(c[1])

    bounds_sw = [min(all_lats), min(all_lngs)]
    bounds_ne = [max(all_lats), max(all_lngs)]

    # ── Load and parse Excel ──────────────────────────────────────────────────
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"Traffic data file not found: {xlsx_path}\n"
            "Please upload Traffic_LOS_Report.xlsx via the dashboard."
        )

    wb = openpyxl.load_workbook(xlsx_path)

    # Try to extract a date from the first sheet (cell A1 or sheet name)
    # Fall back gracefully if not present
    date_label = _extract_date_label(wb)

    all_data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lanes    = ws["C4"].value or 4
        base_cap = ws["C6"].value or 1600
        adj      = ws["C7"].value or 0.85
        capacity = base_cap * lanes * adj

        road_data = []
        for row in ws.iter_rows(min_row=12, max_row=35, values_only=True):
            time_str = row[0]
            if not time_str:
                continue
            car, mc, van, ml, hl, bus = [row[i] or 0 for i in range(1, 7)]
            total = car + mc + van + ml + hl + bus
            pcu   = car*1.0 + mc*0.5 + van*1.5 + ml*2.0 + hl*3.0 + bus*3.0
            vc    = pcu / capacity if capacity > 0 else 0
            road_data.append({
                "time":     time_str,
                "car":      int(car),
                "mc":       int(mc),
                "van":      int(van),
                "ml":       int(ml),
                "hl":       int(hl),
                "bus":      int(bus),
                "total":    int(total),
                "pcu":      round(pcu, 1),
                "capacity": round(capacity, 1),
                "vc":       round(vc, 4),
                "los":      compute_los(vc),
            })
        all_data[sheet] = road_data

    road_count = len(all_data)
    return all_data, road_coords, [bounds_sw, bounds_ne], date_label, road_count


def _extract_date_label(wb):
    """Try to read a date label from the workbook. Returns a fallback string."""
    try:
        ws = wb[wb.sheetnames[0]]
        # Check cell A1 for a date string (adjust if your Excel layout differs)
        val = ws["A1"].value
        if val and isinstance(val, str) and len(val) > 4:
            return val
        # Try sheet name as fallback
        name = wb.sheetnames[0]
        if name and name.lower() not in ("sheet1", "sheet2"):
            return name
    except Exception:
        pass
    return "Traffic LOS Report"
